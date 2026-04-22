#!/usr/bin/env python3
"""Fetch ENARGAS monthly statistics XLSXs.

The "Datos Operativos / Datos Estadísticos" site exposes 10 public .xlsx
files with 30+ years of monthly data. URL pattern:
  https://www.enargas.gob.ar/secciones/transporte-y-distribucion/datos-estadisticos/<CODE>/<FILE>.xlsx
Requires a Referer header; User-Agent alone returns 404.

Wave A only pulls the 3 richest files (cuenca/gasoducto flow + regional
demand + firm contracts). Units in the raw sheets are thousand m³
(dam³) per month. We preserve them verbatim and convert in the frontend.
"""

import io
import json
import os
import sys
from datetime import datetime, timezone

import requests
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _meta import wrap, write_csv  # noqa: E402

OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
BASE = 'https://www.enargas.gob.ar/secciones/transporte-y-distribucion/datos-estadisticos'
REFERER = 'https://www.enargas.gob.ar/secciones/transporte-y-distribucion/datos-operativos.php'
HDRS = {
    'User-Agent': 'Mozilla/5.0 Chrome/130 estado-del-sistema',
    'Referer': REFERER,
}


def fetch_xlsx(code, filename):
    url = f'{BASE}/{code}/{filename}'
    r = requests.get(url, headers=HDRS, timeout=120)
    r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)


def to_iso(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return None


def safe_int(v):
    if v is None or v == '':
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def parse_grt(wb):
    """GRT = Gas Recibido por Transportista. Two sheets: Cuenca, Gasoducto.

    Cuenca headers (rows 12-13):
      TGN (Neuquina, Noroeste, Otros) | TGS (Neuquina, San Jorge, Austral, Otros)
      | Distribuidoras Propios | Otros | Total
    Gasoducto headers (rows 15-16):
      TGN (Centro Oeste, Norte, Otros) | TGS (Neuba I+II, Gral San Martin, Otros)
      | Distribuidoras (Malargüe, Sur, Otros) | Total
    Data units are thousand m³ per month (dam³/month).
    """
    def parse_sheet(s, header_row, cols):
        rows = []
        for i, row in enumerate(s.iter_rows(values_only=True)):
            if i <= header_row:
                continue
            fecha = to_iso(row[0])
            if fecha is None:
                continue
            rec = {'fecha': fecha}
            for key, idx in cols.items():
                if idx < len(row):
                    rec[key] = safe_int(row[idx])
            rows.append(rec)
        return rows

    cuenca_cols = {
        'tgn_neuquina': 1, 'tgn_noroeste': 2, 'tgn_otros': 3,
        'tgs_neuquina': 4, 'tgs_san_jorge': 5, 'tgs_austral': 6, 'tgs_otros': 7,
        'distribuidoras_propios': 8, 'otros_origenes': 9, 'total': 10,
    }
    gas_cols = {
        'tgn_centro_oeste': 1, 'tgn_norte': 2, 'tgn_otros': 3,
        'tgs_neuba': 4, 'tgs_san_martin': 5, 'tgs_otros': 6,
        'distr_malargue': 7, 'distr_sur': 8, 'distr_otros': 9, 'total': 10,
    }
    cuenca = parse_sheet(wb['Cuenca'], header_row=13, cols=cuenca_cols)
    gasoducto = parse_sheet(wb['Gasoducto'], header_row=16, cols=gas_cols)
    return {'cuenca': cuenca, 'gasoducto': gasoducto}


def parse_ged(wb, months_back=60):
    """GED "STS2" sheet = gas delivered, aggregated by distribuidora.

    Columns (r11/r12 header rows):
      1: Ban (Buenos Aires Norte)
      2: Centro
      3-4: Cuyana (Cuyo + Malargue)
      5-7: GasNea (Chaco + Corrientes + Entre Ríos)
      8: Litoral
      9: Metrogas (Metropolitana — includes southern GBA)
      10-11: Noroeste = Gasnor (Salta + Tucumán)
      12-15: Pampeana (Bahía Blanca + Buenos Aires + La Pampa Norte + La Pampa Sur)
      16-20: Sur (Buenos Aires Sur + Chubut Sur + Neuquén + Santa Cruz Sur + Tierra del Fuego)
      21: Total general
    Data starts at r13. Units: miles de m³ de 9300 kcal (dam³/month).
    We keep only the last `months_back` rows to keep the JSON small.
    """
    if 'STS2' not in wb.sheetnames:
        return []
    s = wb['STS2']
    # sum groups: distribuidora_id -> [col indices to sum]
    groups = {
        'naturgy_ban': [1],
        'centro': [2],
        'cuyana': [3, 4],
        'gasnea': [5, 6, 7],
        'litoral': [8],
        'metrogas': [9],
        'gasnor': [10, 11],
        'pampeana': [12, 13, 14, 15],
        'sur': [16, 17, 18, 19, 20],
    }
    rows = []
    for i, row in enumerate(s.iter_rows(values_only=True)):
        if i < 13:
            continue
        fecha = to_iso(row[0])
        if fecha is None:
            continue
        rec = {'fecha': fecha}
        for dist_id, cols in groups.items():
            total = 0
            got_any = False
            for c in cols:
                if c < len(row):
                    v = safe_int(row[c])
                    if v is not None:
                        total += v
                        got_any = True
            rec[dist_id] = total if got_any else None
        rows.append(rec)
    # Keep tail to limit file size. 60 months × 9 ints ≈ 10 KB after compression.
    return rows[-months_back:] if months_back else rows


def parse_contratos(wb):
    """Contratos de transporte firme por licenciataria."""
    # Sheet names vary; take the first non-Indice sheet.
    name = next((n for n in wb.sheetnames if n.lower() != 'indice'), None)
    if not name:
        return []
    s = wb[name]
    # First 15 rows are usually titles/notes; find the first row whose first
    # cell is a date and assume everything above was header.
    rows = list(s.iter_rows(values_only=True))
    header_idx = None
    for i, r in enumerate(rows):
        if isinstance(r[0], datetime):
            header_idx = i
            break
    if header_idx is None:
        return []
    # Header row is right above the first data row. Columns: mostly numbers.
    header = rows[header_idx - 1] if header_idx > 0 else []
    labels = [str(h or '').strip() for h in header]
    out = []
    for r in rows[header_idx:]:
        fecha = to_iso(r[0])
        if fecha is None:
            continue
        rec = {'fecha': fecha}
        for j, v in enumerate(r[1:], start=1):
            key = labels[j] if j < len(labels) else f'col_{j}'
            if not key:
                continue
            val = safe_int(v)
            if val is not None:
                rec[key] = val
        out.append(rec)
    return out


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    payload = {}
    errors = []
    try:
        grt = fetch_xlsx('GRT', 'GRT.xlsx')
        payload['gas_recibido'] = parse_grt(grt)
        print(f"GRT: cuenca={len(payload['gas_recibido']['cuenca'])} rows, "
              f"gasoducto={len(payload['gas_recibido']['gasoducto'])} rows")
    except Exception as e:
        errors.append(f'GRT: {e}')
        print(f'ERROR GRT: {e}', file=sys.stderr)

    try:
        contratos = fetch_xlsx('Contratos', 'Contratos.xlsx')
        payload['contratos_firme'] = parse_contratos(contratos)
        print(f"Contratos: {len(payload['contratos_firme'])} rows")
    except Exception as e:
        errors.append(f'Contratos: {e}')
        print(f'ERROR Contratos: {e}', file=sys.stderr)

    try:
        # GED is 5 MB raw; we only keep the last 60 months × 9 distribuidoras.
        ged = fetch_xlsx('GED', 'GED.xlsx')
        payload['gas_entregado'] = parse_ged(ged, months_back=60)
        print(f"GED: {len(payload['gas_entregado'])} monthly rows (last 60)")
    except Exception as e:
        errors.append(f'GED: {e}')
        print(f'ERROR GED: {e}', file=sys.stderr)

    # Latest fecha across any series for envelope metadata.
    latest = None
    for section in payload.values():
        series = section if isinstance(section, list) else (
            list(section.values())[0] if isinstance(section, dict) else []
        )
        if series:
            d = series[-1].get('fecha')
            if d and (latest is None or d > latest):
                latest = d

    # Trim history to last 15 years for each monthly series — 180 months is
    # more than enough for YoY and cycle views, and keeps the wire size small.
    MONTHS = 180
    if 'gas_recibido' in payload:
        for k in ('cuenca', 'gasoducto'):
            if k in payload['gas_recibido']:
                payload['gas_recibido'][k] = payload['gas_recibido'][k][-MONTHS:]
    if 'contratos_firme' in payload:
        payload['contratos_firme'] = payload['contratos_firme'][-MONTHS:]
    # gas_entregado already trimmed to 60 in parse_ged.

    # Compact JSON (no indent) cuts the file ~3-4x.
    envelope = wrap(payload, source='ENARGAS datos-estadisticos (GRT + Contratos + GED)',
                    source_date=latest, errors=errors or None)
    out_path = os.path.join(OUT_DIR, 'enargas_monthly.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(envelope, f, ensure_ascii=False, separators=(',', ':'))
    print(f"enargas_monthly.json: latest={latest}, size={os.path.getsize(out_path) // 1024} KB")

    # Four sub-CSVs — the JSON is an object with heterogeneous sub-tables so
    # each one becomes its own flat CSV, keeping the audit trail tabular.
    base = os.path.splitext(out_path)[0]
    grt = payload.get('gas_recibido') or {}
    write_csv(f'{base}_cuenca.csv', grt.get('cuenca') or [])
    write_csv(f'{base}_gasoducto.csv', grt.get('gasoducto') or [])
    write_csv(f'{base}_contratos.csv', payload.get('contratos_firme') or [])
    write_csv(f'{base}_entregado.csv', payload.get('gas_entregado') or [])


if __name__ == '__main__':
    main()
