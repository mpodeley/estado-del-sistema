#!/usr/bin/env python3
"""Parse linepack equilibrium Excel into JSON."""

import json
import os
import openpyxl

RAW_DIR = os.path.join(os.path.dirname(__file__), '..', 'raw')
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    # Find the linepack file (name may vary)
    linepack_file = None
    for f in os.listdir(RAW_DIR):
        if 'linepack' in f.lower() and f.endswith('.xlsx'):
            linepack_file = f
            break

    if not linepack_file:
        print("No linepack Excel found, skipping.")
        return

    wb = openpyxl.load_workbook(os.path.join(RAW_DIR, linepack_file), data_only=True)
    ws = wb.active

    # Read headers from row 1
    headers = [cell.value for cell in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [cell.value for cell in ws[r]]
        if vals[0] is None:
            continue
        row = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            v = vals[i] if i < len(vals) else None
            key = h.strip().lower().replace(' ', '_').replace('.', '')
            if isinstance(v, (int, float)):
                row[key] = round(v, 2)
            elif v is not None:
                row[key] = str(v)
        rows.append(row)

    with open(os.path.join(OUT_DIR, 'linepack.json'), 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    print(f"linepack.json: {len(rows)} rows")
    wb.close()


if __name__ == '__main__':
    main()
