#!/usr/bin/env python3
"""Generate a paste-ready 'Datos' sheet for the legacy Excel.

The analyst maintains `Base Reporte Estado de Sistema.xlsx` by hand-typing into
its **"Datos"** sheet ("Conv. valores" is derived from it by formulas). Now that
the pipeline already captures almost everything that sheet needs, we emit a file
whose columns line up 1:1 with "Datos" so the analyst can copy the recent rows
and paste them in instead of typing — keeping the legacy Excel alive without us
touching their file.

Output: public/data/carga_datos.xlsx (+ .csv twin). Columns sit at the exact
"Datos" indices (spacer columns preserved); dates are written so they match the
pre-filled Fecha column. Unmapped columns (Stock GNL, PCS, per-point injection,
tramos-finales limits) stay blank for manual entry.

Column map was read from the real workbook's "Datos" sheet (row 5 = headers,
data from row 6). See docs/datasets or parse_base_excel.py for the sister sheet.
"""

import argparse
import csv
import json
import os
import sys
from datetime import date, datetime, timedelta

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
XLSX_OUT = os.path.join(OUT_DIR, 'carga_datos.xlsx')
CSV_OUT = os.path.join(OUT_DIR, 'carga_datos.csv')

# (col_index_in_Datos, source, field, header_label)
# source: 'date' | 'daily' | 'ps' | 'ppo' | 'wx:<cityid>'
# For 'wx:*', field is one of min/max/prom.
DATOS_COLS = [
    (1, 'date', None, 'Fecha'),
    (2, 'daily', 'demanda_total', 'Demanda DS'),
    (3, 'daily', 'prioritaria', 'Prioritaria'),
    (4, 'daily', 'industria', 'Industria'),
    (5, 'daily', 'usinas', 'Usinas'),
    (6, 'daily', 'exportaciones', 'Exportaciones'),
    (8, 'ps', 'iny_sur', 'Sur'),
    (9, 'ps', 'iny_neuba1', 'NBA I'),
    (10, 'ps', 'iny_neuba2', 'NBA II'),
    (11, 'daily', 'iny_tgs', 'TGS'),
    (13, 'ps', 'iny_norte', 'Norte'),
    (14, 'ps', 'iny_neuquen', 'Nqn'),
    (15, 'daily', 'iny_tgn', 'TGN'),
    (17, 'daily', 'iny_gpm', 'GPM'),
    (18, 'daily', 'iny_bolivia', 'BOLIVIA'),
    (19, 'daily', 'iny_escobar', 'ESCOBAR'),
    (20, 'daily', 'iny_enarsa', 'ENARSA'),
    (22, 'wx:tucuman', 'min', 'Tucumán Min'),
    (23, 'wx:tucuman', 'max', 'Tucumán Max'),
    (24, 'wx:tucuman', 'prom', 'Tucumán Promedio'),
    (25, 'temp_ba', 'min', 'BA Min'),
    (26, 'temp_ba', 'max', 'BA Max'),
    (27, 'temp_ba', 'prom', 'BA Promedio'),
    (28, 'wx:esquel', 'min', 'Esquel Min'),
    (29, 'wx:esquel', 'max', 'Esquel Max'),
    (30, 'wx:esquel', 'prom', 'Esquel Promedio'),
    (32, 'daily', 'linepack_tgs', 'Linepack TGS'),
    (33, 'daily', 'var_linepack_tgs', 'Var. TGS'),
    (34, 'daily', 'lim_inf_tgs', 'Lim Inf. TGS'),
    (35, 'daily', 'lim_sup_tgs', 'Lim Sup. TGS'),
    (37, 'daily', 'linepack_tgn', 'Linepack TGN'),
    (38, 'daily', 'var_linepack_tgn', 'Var. TGN'),
    (39, 'daily', 'lim_inf_tgn', 'Lim Inf. TGN'),
    (40, 'daily', 'lim_sup_tgn', 'Lim Sup. TGN'),
    (42, 'daily', 'tramo_final_tgs', 'Tramo Final TGS'),
    (45, 'daily', 'tramo_final_tgn', 'Tramo Final TGN'),
    (64, 'date', None, 'Fecha'),
    (65, 'ppo', 'gas_mmm3', 'GAS'),
    (66, 'ppo', 'gasoil_m3', 'GASOIL'),
    (67, 'ppo', 'fueloil_tn', 'FUELOIL'),
    (68, 'ppo', 'carbon_tn', 'CARBON'),
]
MAX_COL = max(c for c, *_ in DATOS_COLS)

# Columns whose value is a day-over-day delta of a linepack series, used to fill
# the variation when the source didn't carry it.
VAR_FROM = {'var_linepack_tgs': 'linepack_tgs', 'var_linepack_tgn': 'linepack_tgn'}


def _load(name):
    path = os.path.join(OUT_DIR, name)
    if not os.path.exists(path):
        return []
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)
    return (raw.get('data') if isinstance(raw, dict) else raw) or []


def _by_fecha(rows):
    return {r['fecha']: r for r in rows if r.get('fecha')}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--days', type=int, default=60, help='recent window to emit')
    args = ap.parse_args()

    daily = _by_fecha(_load('daily.json'))
    ps = _by_fecha(_load('enargas_ps.json'))
    ppo = _by_fecha(_load('cammesa_ppo.json'))

    # weather_history: {cityid: {fecha: {min,max,prom}}}
    wx = {}
    for city in _load('weather_history.json'):
        cid = city.get('id')
        if not cid:
            continue
        wx[cid] = {
            h['fecha']: {'min': h.get('temp_min'), 'max': h.get('temp_max'), 'prom': h.get('temp_prom')}
            for h in city.get('history', []) if h.get('fecha')
        }

    if not daily:
        print('ERROR: daily.json empty — nothing to generate', file=sys.stderr)
        return 1

    # Spine = the last N daily dates.
    all_fechas = sorted(daily.keys())
    cutoff = (datetime.strptime(all_fechas[-1], '%Y-%m-%d').date() - timedelta(days=args.days)).isoformat()
    fechas = [f for f in all_fechas if f >= cutoff]

    def value(source, field, fecha):
        if source == 'date':
            return None  # handled separately (written as a real date)
        if source == 'daily':
            return daily.get(fecha, {}).get(field)
        if source == 'ps':
            return ps.get(fecha, {}).get(field)
        if source == 'ppo':
            return ppo.get(fecha, {}).get(field)
        if source == 'temp_ba':
            # RDS (current) first, then the weather archive for the days RDS slims.
            v = daily.get(fecha, {}).get(f'temp_{field}_ba')
            if v is None:
                v = (wx.get('ba', {}).get(fecha) or {}).get(field)
            return v
        if source.startswith('wx:'):
            return (wx.get(source[3:], {}).get(fecha) or {}).get(field)
        return None

    # Build a grid: list of {col: value}. Dates handled per-row.
    grid = []
    for fecha in fechas:
        d = datetime.strptime(fecha, '%Y-%m-%d').date()
        row = {1: d, 64: d}  # Fecha + the fuels sub-block date
        for col, source, field, _label in DATOS_COLS:
            if source == 'date':
                continue
            v = value(source, field, fecha)
            # Redondear floats: algunos campos derivados (ej. exportaciones =
            # exp_tgn+exp_tgs en daily.json) arrastran ruido de punto flotante
            # que ensucia la hoja al pegarla en Excel.
            row[col] = round(v, 2) if isinstance(v, float) else v
        grid.append(row)

    # Fill missing linepack variations from day-over-day diffs.
    col_of = {field: col for col, _s, field, _l in DATOS_COLS}
    prev = {}
    for row in grid:
        for var_field, lp_field in VAR_FROM.items():
            var_col, lp_col = col_of[var_field], col_of[lp_field]
            cur = row.get(lp_col)
            if row.get(var_col) is None and cur is not None and prev.get(lp_field) is not None:
                row[var_col] = round(cur - prev[lp_field], 2)
            if cur is not None:
                prev[lp_field] = cur

    headers = {c: label for c, _s, _f, label in DATOS_COLS}

    # --- XLSX ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Datos (carga)'
    for c in range(1, MAX_COL + 1):
        ws.cell(1, c).value = headers.get(c)
    for i, row in enumerate(grid, start=2):
        for c in range(1, MAX_COL + 1):
            v = row.get(c)
            if v is not None:
                ws.cell(i, c).value = v
        ws.cell(i, 1).number_format = 'yyyy-mm-dd'
        ws.cell(i, 64).number_format = 'yyyy-mm-dd'
    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(XLSX_OUT)

    # --- CSV twin (same column alignment, ISO dates) ---
    with open(CSV_OUT, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow([headers.get(c, '') for c in range(1, MAX_COL + 1)])
        for row in grid:
            out = []
            for c in range(1, MAX_COL + 1):
                v = row.get(c)
                out.append(v.isoformat() if isinstance(v, date) else ('' if v is None else v))
            w.writerow(out)

    print(f"carga_datos.xlsx / .csv: {len(grid)} rows ({fechas[0]} -> {fechas[-1]}), {MAX_COL} cols")
    return 0


if __name__ == '__main__':
    sys.exit(main())
