#!/usr/bin/env python3
"""Parse 'Base Reporte Estado de Sistema.xlsx' into JSON files for the dashboard.

Uses header-driven column mapping: validates row 1 (group) + row 2 (subheader)
match the expected schema and reports loud when anything shifts.
"""

import os
import sys
import unicodedata
from datetime import datetime
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _meta import write_json, write_csv, json_to_csv_path  # noqa: E402

RAW_DIR = os.path.join(os.path.dirname(__file__), '..', 'raw')
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')

# Expected schema for sheet "Conv. valores"
# (group_row1, header_row2, internal_key, treat_zero_as_null)
SCHEMA = [
    (None, 'FECHA', 'fecha', False),
    (None, 'Demanda Total DS', 'demanda_total', True),
    (None, 'Prioritaria', 'prioritaria', True),
    (None, 'Industria', 'industria', False),
    (None, 'Usinas', 'usinas', False),
    (None, 'Exportaciones', 'exportaciones', False),
    (None, 'TGS', 'iny_tgs', False),
    (None, 'TGN', 'iny_tgn', False),
    (None, 'Iny. Enarsa', 'iny_enarsa', False),
    (None, 'GPM', 'iny_gpm', False),
    (None, 'Bolivia', 'iny_bolivia', False),
    (None, 'ESCOBAR', 'iny_escobar', False),
    (None, 'Iny. Total', 'iny_total', False),
    (None, 'Linepack TGS', 'linepack_tgs', True),
    (None, 'Var. Linepack TGS', 'var_linepack_tgs', True),
    (None, 'Limite Inf.', 'lim_inf_tgs', False),
    (None, 'Limite Sup.', 'lim_sup_tgs', False),
    (None, 'Linepack TGN', 'linepack_tgn', True),
    (None, 'Var. Linepack TGN', 'var_linepack_tgn', True),
    (None, 'Limite Inf.', 'lim_inf_tgn', False),
    (None, 'Limite Sup.', 'lim_sup_tgn', False),
    (None, 'Linepack Total', 'linepack_total', True),
    (None, 'Variacion Total', 'var_linepack_total', True),
    (None, 'Limite Inf.', 'lim_inf_total', False),
    (None, 'Limite Sup.', 'lim_sup_total', False),
    ('Tramos finales', 'TGS', 'tramo_final_tgs', False),
    ('Tramos finales', 'Limite Inf.', 'tf_lim_inf_tgs', False),
    ('Tramos finales', 'Limite Sup.', 'tf_lim_sup_tgs', False),
    ('Tramos finales', 'TGN', 'tramo_final_tgn', False),
    ('Tramos finales', 'Limite Inf.', 'tf_lim_inf_tgn', False),
    ('Tramos finales', 'Limite Sup.', 'tf_lim_sup_tgn', False),
    ('BUENOS AIRES', 'Temp. Min', 'temp_min_ba', False),
    ('BUENOS AIRES', 'Temp. Max', 'temp_max_ba', False),
    ('BUENOS AIRES', 'Temp Promedio', 'temp_prom_ba', False),
    ('ESQUEL', 'Temp. Min', 'temp_min_esquel', False),
    ('ESQUEL', 'Temp. Max', 'temp_max_esquel', False),
    ('ESQUEL', 'Temp Promedio', 'temp_prom_esquel', False),
    ('CAMMESA', 'GAS', 'cammesa_gas', False),
    ('CAMMESA', 'GAS OIL', 'cammesa_gasoil', False),
    ('CAMMESA', 'FUEL OIL', 'cammesa_fueloil', False),
    ('CAMMESA', 'CARBON', 'cammesa_carbon', False),
    ('CAMMESA', 'TOTAL', 'cammesa_total', False),
]

# Keys that must parse correctly or the pipeline aborts.
CRITICAL_KEYS = {'fecha', 'demanda_total', 'linepack_tgs', 'linepack_tgn'}

# First data column is col 2 (FECHA) — schema starts aligned there.
FIRST_DATA_COL = 2


def norm(s):
    """Normalize for comparison: strip accents, lowercase, collapse whitespace."""
    if s is None:
        return ''
    s = str(s).strip()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.lower().split())


def safe_float(v):
    if v is None:
        return None
    try:
        return round(float(v), 2)
    except (ValueError, TypeError):
        return None


def read_group_row(ws, row=1):
    """Read row 1 with forward-fill (merged cells return None on non-lead columns)."""
    groups = {}
    last = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row, c).value
        if v is not None:
            last = v
        groups[c] = last
    # Clear forward-fill once a new top-level section starts — detect by the
    # presence of known section names. Here we keep it simple: reset to None
    # when the forward-filled label no longer applies. Since the schema
    # declares expected groups, any stale fill will be caught by validation.
    return groups


def build_column_map(ws):
    """Validate headers match SCHEMA. Returns (col_map, issues).

    col_map: {internal_key: column_number}
    issues: list of strings describing mismatches.
    """
    col_map = {}
    issues = []
    groups = read_group_row(ws)

    for i, (expected_group, expected_header, key, _) in enumerate(SCHEMA):
        col = FIRST_DATA_COL + i
        actual_group = groups.get(col) if expected_group is not None else None
        actual_header = ws.cell(2, col).value

        group_ok = (expected_group is None) or (norm(actual_group) == norm(expected_group))
        header_ok = norm(actual_header) == norm(expected_header)

        if group_ok and header_ok:
            col_map[key] = col
        else:
            issues.append(
                f"col {col} ({key}): expected ({expected_group!r}, {expected_header!r}) "
                f"got ({actual_group!r}, {actual_header!r})"
            )

    return col_map, issues


def parse_conv_valores(wb):
    ws = wb['Conv. valores']
    col_map, issues = build_column_map(ws)

    if issues:
        print("WARN: header mismatches in 'Conv. valores':", file=sys.stderr)
        for msg in issues:
            print(f"  {msg}", file=sys.stderr)

    missing_critical = CRITICAL_KEYS - col_map.keys()
    if missing_critical:
        print(f"ERROR: critical columns missing: {sorted(missing_critical)}", file=sys.stderr)
        sys.exit(2)

    # Which keys should have 0.0 → None
    zero_as_null = {k for _, _, k, z in SCHEMA if z}

    rows = []
    for r in range(3, ws.max_row + 1):
        fecha_val = ws.cell(r, col_map['fecha']).value
        if fecha_val is None:
            continue
        if isinstance(fecha_val, datetime):
            fecha_str = fecha_val.strftime('%Y-%m-%d')
        else:
            fecha_str = str(fecha_val)

        row = {'fecha': fecha_str}
        for key, col in col_map.items():
            if key == 'fecha':
                continue
            v = safe_float(ws.cell(r, col).value)
            if v == 0.0 and key in zero_as_null:
                v = None
            row[key] = v

        # Skip rows where everything important is None (future dates)
        has_data = any(row.get(k) is not None for k in [
            'demanda_total', 'linepack_tgs', 'linepack_tgn', 'cammesa_gas'
        ])
        if not has_data:
            continue

        rows.append(row)

    # Deduplicate by date — keep first occurrence (typically richer)
    seen = set()
    deduped = []
    for row in rows:
        if row['fecha'] in seen:
            continue
        seen.add(row['fecha'])
        deduped.append(row)

    return deduped, len(col_map), len(SCHEMA)


def parse_tramos(wb):
    """Parse transport restrictions from sheet "Datos" cols 82-86.

    Column layout inferred from headers in rows 4-5:
      82: Gas Andes — Autorización (export capacity granted, MMm³/d)
      83: CCO        — Capacidad (pipeline capacity, MMm³/d)
      84: CCO        — Corte (reduction vs capacity; negative = cut)
      85: TGS NQN    — Capacidad
      86: TGS NQN    — Corte
    """
    if 'Datos' not in wb.sheetnames:
        return []
    ws = wb['Datos']
    rows = []
    for r in range(6, ws.max_row + 1):
        fecha = ws.cell(r, 1).value
        if fecha is None:
            continue
        fecha_str = fecha.strftime('%Y-%m-%d') if isinstance(fecha, datetime) else str(fecha)
        gas_andes = safe_float(ws.cell(r, 82).value)
        cco_cap = safe_float(ws.cell(r, 83).value)
        cco_corte = safe_float(ws.cell(r, 84).value)
        tgs_nqn_cap = safe_float(ws.cell(r, 85).value)
        tgs_nqn_corte = safe_float(ws.cell(r, 86).value)
        if all(v is None for v in (gas_andes, cco_cap, cco_corte, tgs_nqn_cap, tgs_nqn_corte)):
            continue
        rows.append({
            'fecha': fecha_str,
            'gas_andes_autorizacion': gas_andes,
            'cco_capacidad': cco_cap,
            'cco_corte': cco_corte,
            'tgs_nqn_capacidad': tgs_nqn_cap,
            'tgs_nqn_corte': tgs_nqn_corte,
        })
    return rows


def parse_comments(wb):
    comments = {}

    for sheet, bucket in [('Diario', 'daily'), ('Proyección Semanal', 'weekly')]:
        if sheet not in wb.sheetnames:
            comments[bucket] = []
            continue
        ws = wb[sheet]
        out = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and len(v) > 20:
                    out.append(v.strip())
        comments[bucket] = out

    return comments


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    xlsx_path = os.path.join(RAW_DIR, 'Base Reporte Estado de Sistema.xlsx')
    if not os.path.exists(xlsx_path):
        print(f"ERROR: {xlsx_path} not found", file=sys.stderr)
        sys.exit(2)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    daily, matched, expected = parse_conv_valores(wb)
    print(f"Column mapping: {matched}/{expected} headers matched")

    latest_date = daily[-1]['fecha'] if daily else None
    source = 'Base Reporte Estado de Sistema.xlsx'

    daily_path = os.path.join(OUT_DIR, 'daily.json')
    write_json(
        daily_path,
        daily, source=source, source_date=latest_date,
        headers_matched=matched, headers_expected=expected,
    )
    # CSV with all schema columns, in schema order (first is 'fecha').
    daily_cols = [key for _, _, key, _ in SCHEMA]
    write_csv(json_to_csv_path(daily_path), daily, fieldnames=daily_cols)
    if daily:
        print(f"daily.json: {len(daily)} rows ({daily[0]['fecha']} to {daily[-1]['fecha']})")
    else:
        print("daily.json: 0 rows - check Excel content", file=sys.stderr)

    # Manual comments from the Excel — auto-generator in generate_forecast.py
    # overlays the authoritative comments.json; we save the manual ones separately
    # so they can be surfaced in the UI if ever needed.
    manual = parse_comments(wb)
    manual_path = os.path.join(OUT_DIR, 'comments_manual.json')
    write_json(
        manual_path,
        manual, source=source, source_date=latest_date,
    )
    manual_rows = (
        [{'tipo': 'daily', 'texto': t} for t in manual.get('daily', [])]
        + [{'tipo': 'weekly', 'texto': t} for t in manual.get('weekly', [])]
    )
    write_csv(json_to_csv_path(manual_path), manual_rows, fieldnames=['tipo', 'texto'])
    print(f"comments_manual.json: {len(manual.get('daily', []))} daily, {len(manual.get('weekly', []))} weekly")

    # Transport restrictions: Gas Andes / CCO / TGS NQN from sheet "Datos".
    tramos = parse_tramos(wb)
    tramos_latest = tramos[-1]['fecha'] if tramos else None
    tramos_path = os.path.join(OUT_DIR, 'tramos.json')
    write_json(tramos_path, tramos, source=source, source_date=tramos_latest)
    write_csv(json_to_csv_path(tramos_path), tramos)
    print(f"tramos.json: {len(tramos)} rows")

    wb.close()


if __name__ == '__main__':
    main()
