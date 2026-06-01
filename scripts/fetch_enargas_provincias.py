#!/usr/bin/env python3
"""Fetch ENARGAS gas-entregado por provincia (monthly).

Source: ENARGAS "Datos Estadísticos" → GED.xlsx, sheet "PTS"
  = cuadro 1.06 "Gas Entregado y Nro. Usuarios, por Tipo de Servicio y Provincia".
URL pattern + Referer header identical to fetch_enargas_estadisticas.py.

The PTS sheet is a pre-rendered pivot: columns are distribuidora→provincia, rows
are months, the value is gas entregado (miles de m³ de 9300 kcal = dam³/mes), with
the *tipo de servicio* dimension collapsed. A single province (Buenos Aires) shows
up under several distribuidoras, so we sum every column that maps to the same
province slug to get a per-province total.

Why only the province total and not the segment split: ENARGAS publishes the
tipo-de-servicio breakdown only collapsed to the national level (sheet STS); no
public flat table crosses provincia × segmento. So this feeds the choropleth
(consumo total / km²) and the per-province monthly trend — not a per-province
segment donut.
"""

import io
import os
import sys
import unicodedata
from datetime import datetime

import requests
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _meta import write_json, write_csv, json_to_csv_path  # noqa: E402

OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
URL = ('https://www.enargas.gob.ar/secciones/transporte-y-distribucion/'
       'datos-estadisticos/GED/GED.xlsx')
REFERER = ('https://www.enargas.gob.ar/secciones/transporte-y-distribucion/'
           'datos-operativos.php')
HDRS = {
    'User-Agent': 'Mozilla/5.0 Chrome/130 estado-del-sistema',
    'Referer': REFERER,
}

# Keep ~10 years of monthly history — enough for the per-province trend chart
# while keeping the JSON small.
MONTHS_BACK = 120

# Row layout of the PTS sheet (0-based, matches iter_rows order):
#   11 = distribuidora header, 12 = provincia header, 13.. = monthly data.
ROW_DIST = 11
ROW_PROV = 12
ROW_DATA0 = 13

# Pretty display name per slug (for the sibling geojson + UI). Spanish accents.
PROV_DISPLAY = {
    'buenos_aires': 'Buenos Aires',
    'caba': 'CABA',
    'catamarca': 'Catamarca',
    'cordoba': 'Córdoba',
    'la_rioja': 'La Rioja',
    'mendoza': 'Mendoza',
    'san_juan': 'San Juan',
    'san_luis': 'San Luis',
    'chaco': 'Chaco',
    'corrientes': 'Corrientes',
    'entre_rios': 'Entre Ríos',
    'santa_fe': 'Santa Fe',
    'jujuy': 'Jujuy',
    'salta': 'Salta',
    'santiago_del_estero': 'Santiago del Estero',
    'tucuman': 'Tucumán',
    'la_pampa': 'La Pampa',
    'chubut': 'Chubut',
    'neuquen': 'Neuquén',
    'rio_negro': 'Río Negro',
    'santa_cruz': 'Santa Cruz',
    'tierra_del_fuego': 'Tierra del Fuego',
}


def province_slug(name):
    """Normalise a province label to a stable join key.

    Generic: strip accents, lowercase, non-alphanumerics → underscore.
    Special case: ENARGAS labels CABA as 'Capital Federal'.
    """
    s = (name or '').strip()
    if not s:
        return None
    norm = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    norm = norm.lower().strip()
    if norm in ('capital federal', 'ciudad de buenos aires', 'caba'):
        return 'caba'
    out = []
    prev_us = False
    for ch in norm:
        if ch.isalnum():
            out.append(ch)
            prev_us = False
        elif not prev_us:
            out.append('_')
            prev_us = True
    return ''.join(out).strip('_')


def to_iso(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return None


def safe_num(v):
    if v is None or v == '':
        return None
    try:
        return round(float(v), 1)
    except (TypeError, ValueError):
        return None


def fetch_pts():
    r = requests.get(URL, headers=HDRS, timeout=120)
    r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    if 'PTS' not in wb.sheetnames:
        raise RuntimeError(f"GED.xlsx has no 'PTS' sheet (sheets: {wb.sheetnames})")
    return wb['PTS']


def parse_pts(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    prov_header = rows[ROW_PROV]
    dist_header = rows[ROW_DIST]

    # Map column index -> province slug, skipping the 'Total general' column
    # (it has a distribuidora-row label but no province-row label).
    col_slug = {}
    slugs_seen = []
    for idx, raw in enumerate(prov_header):
        if idx == 0:
            continue  # date column
        dist = str(dist_header[idx] or '').strip().lower() if idx < len(dist_header) else ''
        if 'total' in dist:
            continue
        slug = province_slug(raw)
        if not slug:
            continue
        col_slug[idx] = slug
        if slug not in slugs_seen:
            slugs_seen.append(slug)

    out = []
    for r in rows[ROW_DATA0:]:
        fecha = to_iso(r[0])
        if fecha is None:
            continue
        rec = {'fecha': fecha}
        acc = {}
        for idx, slug in col_slug.items():
            if idx >= len(r):
                continue
            v = safe_num(r[idx])
            if v is None:
                continue
            acc[slug] = (acc.get(slug) or 0) + v
        for slug in slugs_seen:
            rec[slug] = round(acc[slug], 1) if slug in acc else None
        out.append(rec)
    return out[-MONTHS_BACK:] if MONTHS_BACK else out, slugs_seen


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    sheet = fetch_pts()
    data, slugs = parse_pts(sheet)
    latest = data[-1]['fecha'] if data else None
    print(f"PTS: {len(data)} monthly rows, {len(slugs)} provincias, latest={latest}")
    print(f"  provincias: {', '.join(slugs)}")

    out_path = os.path.join(OUT_DIR, 'enargas_provincias.json')
    write_json(
        out_path, data,
        source='ENARGAS datos-estadisticos (GED.xlsx / cuadro 1.06 PTS)',
        source_date=latest,
        units='miles de m³ de 9300 kcal por mes (dam³/mes)',
        provincias=slugs,
    )
    write_csv(json_to_csv_path(out_path), data, fieldnames=['fecha'] + slugs)
    print(f"enargas_provincias.json: {os.path.getsize(out_path) // 1024} KB")


if __name__ == '__main__':
    main()
